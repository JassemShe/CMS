from tkinter import *
from tkinter import messagebox
from tkinter import ttk
import mysql.connector
import openpyxl


def main_gui():
    """This the main window of the program."""
    import register

    root = Tk()
    root.title("Cafe System Managment")
    root.geometry("320x280")

    # Create a mysql instant and connect it (will add database later when create it)
    mydb = mysql.connector.connect(
                      host = "127.0.0.1",
                      user = "root",
                      passwd = "jmjmaaskull",
                      database = "cms")

    # Create cursor object
    cursor = mydb.cursor()


    def login_fun():
        """This function will open a new window if the username and password are valid.

        The window will inculde the options the user can use in the application"""
        global username_try

        login = Tk()
        login.title("Option menu")
        login.geometry("320x340")

        def sales():
            """This function will open sales window to apply sales operations"""
            sales = Tk()
            sales.title("Sales")
            sales.geometry("420x400")


            #Totol price
            global total_price
            total_price = 0
            def add_items():
                """add chosen items to total_price"""
                global total_price
                item = float(item_drop.get())
                quantiy = int(item_quantity_box.get())
                price = item * quantiy
                total_price = total_price + price
                total_price_value_label["text"] = round(total_price,2)
                item_quantity_box.delete(0,END)

            def add_tea():
                """add chosen items to total_price"""
                global total_price
                tea = float(tea_drop.get())
                quantiy = int(tea_quantity_box.get())
                price = tea * quantiy
                total_price = total_price + price
                total_price_value_label["text"] = round(total_price,2)
                tea_quantity_box.delete(0,END)

            def add_coffe():
                """add chosen items to total_price"""
                global total_price
                coffe = float(coffe_drop.get())
                quantiy = int(coffe_quantity_box.get())
                price = coffe * quantiy
                total_price = total_price + price
                total_price_value_label["text"] = round(total_price,2)
                coffe_quantity_box.delete(0,END)

            def paid():
               """Calcualte remenaning amount for customer"""
               global total_price
               paid = float(paid_box.get())
               remaning = paid - total_price
               remaining_value_label["text"] = round(remaning,2)
               paid_box.delete(0,END)

            def remove():
               """Reset total price"""
               global total_price
               total_price = 0
               total_price_value_label["text"] = "**"
               remaining_value_label["text"] = "**"

            def enter():
               """enter sales as succesful to database (not finished)"""
               global total_price

               total_price = 0
               total_price_value_label["text"] = "**"
               remaining_value_label["text"] = "**"

            # Add items
            items_label = Label(sales,text = "Items",font=("Arial",12))
            items_label.grid(row = 0 , column = 0 ,pady = (10,0),padx = (0,5),sticky=W)
            items_list = ["0.05","0.10","0.15","0.20","0.25","0.30","0.35","0.40","0.45","0.50","0.55","0.60","0.65","0.70","0.75","0.80","0.85",\
                          "0.90","0.95","1.00"]
            item_drop = ttk.Combobox(sales,values = items_list,state="readonly")
            item_drop.current(0)
            item_drop.grid(row = 1, column = 0 , pady = (0,0))
            item_quantity_label = Label(sales,text = "Q",font=("Arial",12))
            item_quantity_label.grid(row = 0 ,column = 1 ,pady = (15,0) ,padx=(10 , 5))
            item_quantity_box = Entry(sales,width = 10)
            item_quantity_box.grid(row = 1, column = 1 , pady = (0,0), padx =(10,5))
            add_item_button = Button(sales,text = "Add",font=("Arial",10),command=add_items)
            add_item_button.grid(row = 1 , column = 2 , pady = (0,0), padx = (10 ,5))

            # Add tea
            tea_cup_label = Label(sales,text = "Tea cup price",font=("Arial",12))
            tea_cup_label.grid(row = 2 , column = 0 ,pady = (20,0),padx = (0,5),sticky=W)
            tea_list = ["0.25","0.30","0.35","0.40","0.45","0.50"]
            tea_drop = ttk.Combobox(sales,values = tea_list,state="readonly")
            tea_drop.current(2)
            tea_drop.grid(row = 3, column = 0 , pady = (0,0))
            tea_quantity_label = Label(sales,text = "Q",font=("Arial",12))
            tea_quantity_label.grid(row = 2 ,column = 1 ,pady = (15,0) ,padx=(10 , 5))
            tea_quantity_box = Entry(sales,width = 10)
            tea_quantity_box.grid(row = 3, column = 1 , pady = (0,0), padx =(10,5))
            add_tea_button = Button(sales,text = "Add",font=("Arial",10),command=add_tea)
            add_tea_button.grid(row = 3 , column = 2 , pady = (0,0), padx = (10 ,5))

            # Add coffe
            coffe_cup_label = Label(sales,text = "Coffe cup price",font=("Arial",12))
            coffe_cup_label.grid(row = 4 , column = 0 ,pady = (20,0),padx = (0,5),sticky=W)
            coffe_list = ["0.35","0.40","0.45","0.50","0.55","0.60","0.65","0.70","0.80","0.85","0.90","0.95","1.00"]
            coffe_drop = ttk.Combobox(sales,values = coffe_list,state="readonly")
            coffe_drop.current(3)
            coffe_drop.grid(row = 5, column = 0 , pady = (0,0))
            coffe_quantity_label = Label(sales,text = "Q",font=("Arial",12))
            coffe_quantity_label.grid(row = 4 ,column = 1 ,pady = (15,0) ,padx=(10 , 5))
            coffe_quantity_box = Entry(sales,width = 10)
            coffe_quantity_box.grid(row = 5, column = 1 , pady = (0,0), padx =(10,5))
            add_coffe_button = Button(sales,text = "Add",font=("Arial",10),command=add_coffe)
            add_coffe_button.grid(row = 5 , column = 2 , pady = (0,0), padx = (10 ,5))

            # Total price
            total_price_text_label = Label(sales,text="Total Price: ",font=("Arial",14))
            total_price_text_label.grid(row = 6 , column = 0,pady=(40,0))
            total_price_value_label = Label(sales,text="**",font=("Arial",14))
            total_price_value_label.grid(row = 6 , column = 1,pady=(40,0))

            # Paid
            paid_label = Label(sales,text="Paid: ",font=("Arial",14))
            paid_label.grid(row=7,column=0)
            paid_box = Entry(sales,width = 10)
            paid_box.grid(row = 7 , column =1)
            paid_button = Button(sales,text="Paid",font=("Arial",10),command=paid)
            paid_button.grid(row=7,column=2)

            # Remaining
            remaining_text_label = Label(sales,text="Remaining: ",font=("Arial",14))
            remaining_text_label.grid(row= 8, column = 0)
            remaining_value_label = Label(sales,text="**",font=("Arial",14))
            remaining_value_label.grid(row=8,column=1)

            # Enter
            enter_button = Button(sales,text="Enter",font=("Arial",10),width = 15)
            enter_button.grid(row=9,column=0,pady=(10,0))

            # Remove
            remove_button = Button(sales,text="Remove",font=("Arial",10),width = 15,command=remove)
            remove_button.grid(row = 9 , column =1 ,pady=(10,0))

        def store():
            """This function will open store window to add items"""
            store = Tk()
            store.title("Store")
            store.geometry("340x400")

            def add_items():
               """Add selected item quantity to items table"""
               selected = item_drop.get().replace(".","_")
               try: # Make sure valid integer value entred
                amount = int(item_quantity_box.get())
                cursor.execute(f"SELECT {selected} FROM items WHERE item_id = 1")
                current_amount = cursor.fetchall()[0][0]
                if current_amount == None:
                    current_amount = 0
                new_amount = current_amount + amount
                cursor.execute(f"UPDATE items SET {selected} = {new_amount} WHERE item_id = 1")
                mydb.commit()
                item_quantity_box.delete(0,END)
               except:
                messagebox.showerror(title="Warning",message="Enter amout as numbers")

            def add_tea():
               """Add selected item quantity to items table"""
               selected = tea_drop.get().replace(".","_")
               try: # Make sure valid integer value entred
                amount = int(tea_quantity_box.get())
                cursor.execute(f"SELECT {selected} FROM tea_cup WHERE tea_id = 1")
                current_amount = cursor.fetchall()[0][0]
                if current_amount == None:
                    current_amount = 0
                new_amount = current_amount + amount
                cursor.execute(f"UPDATE tea_cup SET {selected} = {new_amount} WHERE tea_id = 1")
                mydb.commit()
                tea_quantity_box.delete(0,END)
               except:
                messagebox.showerror(title="Warning",message="Enter amout as numbers")

            def add_coffe():
               """Add selected item quantity to items table"""
               selected = coffe_drop.get().replace(".","_")
               try: # Make sure valid integer value entred
                amount = int(coffe_quantity_box.get())
                cursor.execute(f"SELECT {selected} FROM coffe_cup WHERE coffe_id = 1")
                current_amount = cursor.fetchall()[0][0]
                if current_amount == None:
                    current_amount = 0
                new_amount = current_amount + amount
                cursor.execute(f"UPDATE coffe_cup SET {selected} = {new_amount} WHERE coffe_id = 1")
                mydb.commit()
                coffe_quantity_box.delete(0,END)
               except:
                messagebox.showerror(title="Warning",message="Enter amout as numbers")

            def add_cash():
                """Add cash to current cash"""
                try:
                    new_cash = float(cash_box.get())
                    cursor.execute(f"SELECT current_cash FROM cash WHERE cash_id = 1")
                    current_cash = cursor.fetchall()[0][0]
                    if current_cash == None:
                        current_cash = 0
                    else:
                        current_cash = float(current_cash) # Convert decimal to float
                    current_cash = current_cash + new_cash
                    cursor.execute(f"UPDATE cash SET current_cash = {current_cash} WHERE cash_id = 1")
                    mydb.commit()
                    cash_box.delete(0,END)
                except:
                    messagebox.showerror(title="Warning",message="Enter amout as numbers")

            def export():
                """This function will fetch data in store to a xlsx file"""
                data_to_export = export_drop.get()

                # Save items database
                if data_to_export == "Items":
                    try:
                        excel = openpyxl.Workbook()
                        sheet = excel.active
                        sheet.title = "Items"
                        items_list = ["0.05","0.10","0.15","0.20","0.25","0.30","0.35","0.40","0.45","0.50","0.55","0.60","0.65","0.70","0.75","0.80","0.85",\
                              "0.90","0.95","1.00"]
                        sheet.append(items_list)

                        cursor.execute("SELECT * FROM items")
                        result = cursor.fetchall()
                        record_list = []
                        for record in result[0]:
                            if record == None:
                                record = 0
                            record_list.append(record)
                        sheet.append(record_list)
                        excel.save("items.xlsx")
                        messagebox.showinfo(title="Warning",message="Saved")
                    except:
                        messagebox.showerror(title="Warning",message="Something went wrong!")

                # Save tea database
                if data_to_export == "Tea":
                    try:
                        excel = openpyxl.Workbook()
                        sheet = excel.active
                        sheet.title = "Tea"
                        tea_list = ["row","0.25","0.30","0.35","0.40","0.45","0.50"]
                        sheet.append(tea_list)

                        cursor.execute("SELECT * FROM tea_cup")
                        result = cursor.fetchall()
                        record_list = []
                        for record in result[0]:
                            if record == None:
                                record = 0
                            record_list.append(record)
                        sheet.append(record_list)
                        excel.save("tea.xlsx")
                        messagebox.showinfo(title="Warning",message="Saved")
                    except:
                        messagebox.showerror(title="Warning",message="Something went wrong!")

                # Save coffe database
                if data_to_export == "Coffe":
                    try:
                        excel = openpyxl.Workbook()
                        sheet = excel.active
                        sheet.title = "Coffe"
                        coffe_list = ["row","0.35","0.40","0.45","0.50","0.55","0.60","0.65","0.70","0.80","0.85","0.90","0.95","1.00"]
                        sheet.append(coffe_list)

                        cursor.execute("SELECT * FROM coffe_cup")
                        result = cursor.fetchall()
                        record_list = []
                        for record in result[0]:
                            if record == None:
                                record = 0
                            record_list.append(record)
                        sheet.append(record_list)
                        excel.save("coffe.xlsx")
                        messagebox.showinfo(title="Warning",message="Saved")
                    except:
                        messagebox.showerror(title="Warning",message="Something went wrong!")

                # Save cash database
                if data_to_export == "Cash":
                    try:
                        excel = openpyxl.Workbook()
                        sheet = excel.active
                        sheet.title = "Cash"
                        cash_list = ["row","Total Cash"]
                        sheet.append(cash_list)

                        cursor.execute("SELECT * FROM cash")
                        result = cursor.fetchall()
                        record_list = []
                        for record in result[0]:
                            if record == None:
                                record = 0
                            record_list.append(record)
                        sheet.append(record_list)
                        excel.save("cash.xlsx")
                        messagebox.showinfo(title="Warning",message="Saved")
                    except:
                        messagebox.showerror(title="Warning",message="Something went wrong!")

            # Add items
            items_label = Label(store,text = "Items",font=("Arial",12))
            items_label.grid(row = 0 , column = 0 ,pady = (10,0),padx = (0,5),sticky=W)

            items_list = ["0.05","0.10","0.15","0.20","0.25","0.30","0.35","0.40","0.45","0.50","0.55","0.60","0.65","0.70","0.75","0.80","0.85",\
                          "0.90","0.95","1.00"]
            item_drop = ttk.Combobox(store,values = items_list,state="readonly")
            item_drop.current(0)
            item_drop.grid(row = 1, column = 0 , pady = (0,0))

            item_quantity_label = Label(store,text = "Q",font=("Arial",12))
            item_quantity_label.grid(row = 0 ,column = 1 ,pady = (15,0) ,padx=(10 , 5))

            item_quantity_box = Entry(store,width = 10)
            item_quantity_box.grid(row = 1, column = 1 , pady = (0,0), padx =(10,5))

            add_item_button = Button(store,text = "Add items",font=("Arial",10),command=add_items)
            add_item_button.grid(row = 1 , column = 2 , pady = (0,0), padx = (10 ,5))

            # Add tea cups
            tea_cup_label = Label(store,text = "Tea cup price",font=("Arial",12))
            tea_cup_label.grid(row = 2 , column = 0 ,pady = (20,0),padx = (0,5),sticky=W)

            tea_list = ["0.25","0.30","0.35","0.40","0.45","0.50"]
            tea_drop = ttk.Combobox(store,values = tea_list,state="readonly")
            tea_drop.current(2)
            tea_drop.grid(row = 3, column = 0 , pady = (0,0))

            tea_quantity_label = Label(store,text = "Q",font=("Arial",12))
            tea_quantity_label.grid(row = 2 ,column = 1 ,pady = (15,0) ,padx=(10 , 5))

            tea_quantity_box = Entry(store,width = 10)
            tea_quantity_box.grid(row = 3, column = 1 , pady = (0,0), padx =(10,5))

            add_tea_button = Button(store,text = "Add tea cups",font=("Arial",10),command=add_tea)
            add_tea_button.grid(row = 3 , column = 2 , pady = (0,0), padx = (10 ,5))

            # Add coffe cup
            coffe_cup_label = Label(store,text = "Coffe cup price",font=("Arial",12))
            coffe_cup_label.grid(row = 4 , column = 0 ,pady = (20,0),padx = (0,5),sticky=W)

            coffe_list = ["0.35","0.40","0.45","0.50","0.55","0.60","0.65","0.70","0.80","0.85","0.90","0.95","1.00"]
            coffe_drop = ttk.Combobox(store,values = coffe_list,state="readonly")
            coffe_drop.current(3)
            coffe_drop.grid(row = 5, column = 0 , pady = (0,0))

            coffe_quantity_label = Label(store,text = "Q",font=("Arial",12))
            coffe_quantity_label.grid(row = 4 ,column = 1 ,pady = (15,0) ,padx=(10 , 5))

            coffe_quantity_box = Entry(store,width = 10)
            coffe_quantity_box.grid(row = 5, column = 1 , pady = (0,0), padx =(10,5))

            add_coffe_button = Button(store,text = "Add coffe cups",font=("Arial",10),command=add_coffe)
            add_coffe_button.grid(row = 5 , column = 2 , pady = (0,0), padx = (10 ,5))

            # Add cash
            cash_label = Label(store,text = "Cash",font=("Arial",12))
            cash_label.grid(row = 6 , column = 0 ,pady = (50,0),padx = (0,0))

            cash_box = Entry(store,width = 10)
            cash_box.grid(row = 6, column = 1 , pady = (50,0),padx=(0,0))

            add_cash_button = Button(store,text = "Add cash",font=("Arial",10),command=add_cash)
            add_cash_button.grid(row = 6 , column = 2,padx = (0,0),pady=(50,0))

            # Export button
            export_list = ["Items","Tea","Coffe","Cash"]
            export_drop = ttk.Combobox(store,values = export_list,state="readonly")
            export_drop.grid(row = 7 , column = 0,padx = (0,0),pady=(50,0),sticky = W)
            export_drop.current(0)

            export_button = Button(store,text = "Export",font=("Arial",10),command=export)
            export_button.grid(row = 8 , column = 0,padx = (0,0),pady=(5,0),sticky = W)

        def log_out():
            """This funtion will close login window and open mainGUI window"""
            login.destroy()
            main_gui()

        # Create user name label to show which user is using the program currenty
        user_name_label = Label(login,text =username_try)
        user_name_label.grid(column =  0 ,row = 0,padx=5,pady=5)

        # Create Sales operations button
        sales_button = Button(login, text = " Sales operations ",font=("Arial",16),command=sales)
        sales_button.grid(column = 1, row =1,pady=(60,10))

        # Create Store button
        store_button = Button(login, text = "Store",font=("Arial",16),width=15,command=store)
        store_button.grid(column = 1, row = 2,pady=(10,20))

        # Create log out button
        logout_button = Button(login, text = "Log out",font=("Arial",16),width=15,command=log_out)
        logout_button.grid(column = 1, row = 3)

    def login_valid():
        """This funtion is responible to check if username and password are correct and if correct
        will call login_fun to enter the option menu"""
        global username_try

        username_try = username_box.get()
        password_try = password_box.get()

        # Get users in data
        cursor.execute("SELECT username FROM users")
        users_result = cursor.fetchall()
        user_records = [record[0] for record in users_result]

        # Check if username is in data
        if username_try in user_records:
            # Check if the password match
            sql_command = "SELECT password FROM users WHERE username IN (%s)"
            value = (username_try,)
            cursor.execute(sql_command,value)
            password_result = cursor.fetchall()
            password_records = [record[0] for record in password_result]
            if password_try in password_records:
                root.destroy()
                login_fun()
            else:
                messagebox.showerror(title="Warning",message="Wrong Password")
        else:
            messagebox.showerror(title="Warning",message="Wrong Username")


    # Cafe name title label
    cafe_name_label = Label(root,text="Cafe Managment System",font=("Arial",16)).grid(row = 0 , column = 0,columnspan=2,sticky=W)

    # Create username and password labels
    username_label = Label(root,text="Username",font=("Arial",12)).grid(row = 1, column = 0,pady=(50,0),padx=(40,0))
    password_label = Label(root,text="Password",font=("Arial",12)).grid(row = 2, column = 0,pady=(20,0),padx=(40,0))

    # Create entry boxes for username and passwosrd
    username_box = Entry(root)
    username_box.grid(row = 1,column =1,pady=(60,0),padx=(5,0))
    password_box = Entry(root,show="*")
    password_box.grid(row = 2,column =1,pady=(20,0),padx=(5,0))

    # Create login button
    login_button = Button(root,text="Login",font=("Arial",12),command=login_valid)
    login_button.grid(row = 3,column =1,pady=(20,0),padx=(5,0))

    # Create register button
    register_button = Button(root,text="Register",command=register.register_fun)
    register_button.grid(row = 4,column =1,pady=(20,0),padx=(5,0))

    root.mainloop()

main_gui()
